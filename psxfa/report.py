"""Run a ticker through the engine and render it to console or an XLSX workbook."""
from __future__ import annotations
import argparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .config import DEFAULT_RISKFREE, PILLARS, BANK_PILLARS
from .data import build_financials
from .metrics import cross_check
from .scoring import evaluate, evaluate_bank
from .flags import compute_flags, human_checklist, beneish_partial

FILL = {"DANGER": "FFF8CBAD", "WARN": "FFFFE699", "INFO": "FFD9D9D9",
        "GOOD": "FFC6E0B4", "HEAD": "FF305496"}
SEV_TAG = {"DANGER": "[DANGER]", "WARN": "[WARN] ", "INFO": "[info] "}


def analyze(sym, riskfree=DEFAULT_RISKFREE):
    """Full analysis bundle: scores, gates, flags and checklist for one ticker."""
    fin = build_financials(sym)
    if fin["is_bank"]:
        r = evaluate_bank(fin, riskfree)
        r["profile"], r["pillar_names"] = "bank", list(BANK_PILLARS)
    else:
        r = evaluate(fin, riskfree)
        r["profile"], r["pillar_names"] = "standard", list(PILLARS)
    r["fin"] = fin
    r["flags"] = compute_flags(r["m"], fin)
    r["checklist"] = human_checklist(r["m"], fin)
    r["beneish_partial"] = beneish_partial(fin, r["m"].get("OCF"))
    return r


def gate_status(r):
    stage0 = "PASS" if not r["stage0"] else "FAIL: " + "; ".join(r["stage0"])
    if r["profile"] == "bank":
        cap = ("FAIL: " + "; ".join(r["cap_fail"])) if r["cap_fail"] else \
              ("UNVERIFIED" if r["cap_unverified"] else "PASS")
        return {"Stage0": stage0, "Safety/Capital": cap, "Cash": "N/A (bank)"}
    safety = ("FAIL: " + "; ".join(r["safety_fail"])) if r["safety_fail"] else \
             ("UNVERIFIED" if r["safety_unverified"] else "PASS")
    cash = ("FAIL: " + "; ".join(r["cash_fail"])) if r["cash_fail"] else \
           ("UNVERIFIED" if r["cash_unverified"] else "PASS")
    return {"Stage0": stage0, "Safety/Capital": safety, "Cash": cash}


def tier_only(verdict):
    return verdict.split(" (")[0].split(" -")[0]


def print_console(r):
    fin, m = r["fin"], r["m"]
    prof = "bank" if r["profile"] == "bank" else "standard"
    print("\n" + "=" * 88)
    print(f" {fin['sym']}   FY{m['year']}   [{prof}]   sector={fin.get('sector')}   "
          f"price={m['price']}")
    print("=" * 88)
    g = gate_status(r)
    print(f"  GATES   : Stage0 {g['Stage0']}  |  Safety/Capital {g['Safety/Capital']}"
          f"  |  Cash {g['Cash']}")
    sc = r["scores"]
    print("  PILLARS : " + "  ".join(
        f"{p} {sc[p]:.1f}" if sc[p] is not None else f"{p} n/a" for p in r["pillar_names"]))
    comp = f"{r['composite']:.2f}/5" if r["composite"] is not None else "n/a"
    print(f"  SCORE   : {comp}   Confidence {r['confidence']} "
          f"(coverage {r['coverage']*100:.0f}%)")
    print(f"  QUANT   : {r['verdict']}")

    print("  --- flags ---")
    for f in r["flags"]:
        print(f"    {SEV_TAG[f.severity]} {f.name}: {f.explanation}")
    if not r["flags"]:
        print("    (none)")

    print("  --- human checklist (printed unchecked) ---")
    for title, _, items in r["checklist"]:
        print(f"    {title}")
        for it in items:
            print(f"       [ ] {it}")
    print(f"  FINAL   : pending human governance review "
          f"(resolves to '{tier_only(r['verdict'])}' once gates + governance confirmed)")


def _verdict_row(r):
    fin, m = r["fin"], r["m"]
    g = gate_status(r)
    sc = r["scores"]
    row = {
        "Symbol": fin["sym"], "Sector": fin.get("sector"), "Profile": r["profile"],
        "FY": m.get("year"), "QuantVerdict": r["verdict"],
        "FinalStatus": "PENDING - human governance review",
        "Composite": round(r["composite"], 2) if r["composite"] is not None else None,
        "Confidence": r["confidence"], "Stage0": g["Stage0"],
        "Safety/Capital": g["Safety/Capital"], "Cash": g["Cash"],
    }
    for p in ("Profitability", "Safety", "Cash", "Capital", "Growth", "Valuation"):
        row[f"Score_{p}"] = round(sc[p], 2) if (p in sc and sc[p] is not None) else None
    for k in ("ROE_%", "ROA_%", "GrossMargin_%", "OperatingMargin_%", "NetMargin_%",
              "EBITDAMargin_%", "ROCE_%", "CurrentRatio", "QuickRatio", "DebtToEquity",
              "LT_DebtToEquity", "DebtToEBITDA", "InterestCover", "OCF_to_NI",
              "cumOCF_to_NI", "FCF_margin_%", "PE", "PB", "EarningsYield_%",
              "DividendYield_%", "PayoutRatio_%", "RevenueCAGR_%", "EPS_CAGR_%",
              "BVPS_CAGR_%", "MarginTrend_pp", "AltmanZ", "NetInterestMargin_%",
              "CapitalAdequacy_proxy_%", "ADR_%", "InvestmentToDeposit_%"):
        v = m.get(k)
        row[k] = round(v, 2) if isinstance(v, float) else v
    row["BeneishPartial"] = r["beneish_partial"]
    row["Flags"] = " | ".join(f"{SEV_TAG[f.severity].strip()} {f.name}" for f in r["flags"])
    return row


def _style_header(ws):
    head = PatternFill("solid", fgColor=FILL["HEAD"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = head
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def write_xlsx(results, path):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Verdicts"
    rows = [_verdict_row(r) for r in results]
    cols = list(rows[0].keys())
    ws.append(cols)
    tier_fill = {"STRONG BUY": FILL["GOOD"], "BUY": FILL["GOOD"],
                 "HOLD": FILL["WARN"], "AVOID": FILL["DANGER"], "INSUFFICIENT": FILL["INFO"]}
    vcol = cols.index("QuantVerdict") + 1
    for row in rows:
        ws.append([row[c] for c in cols])
        key = next((k for k in tier_fill if row["QuantVerdict"].startswith(k.split()[0])), None)
        if key:
            ws.cell(ws.max_row, vcol).fill = PatternFill("solid", fgColor=tier_fill[key])
    _style_header(ws)
    ws.column_dimensions["A"].width = 10
    ws.auto_filter.ref = ws.dimensions

    wf = wb.create_sheet("Flags")
    wf.append(["Symbol", "Severity", "Flag", "Value", "Explanation"])
    for r in results:
        for f in r["flags"]:
            wf.append([r["fin"]["sym"], f.severity, f.name, f.value, f.explanation])
            wf.cell(wf.max_row, 2).fill = PatternFill("solid", fgColor=FILL.get(f.severity, "FFFFFFFF"))
    _style_header(wf)
    for col, w in {"A": 10, "B": 11, "C": 26, "D": 10, "E": 70}.items():
        wf.column_dimensions[col].width = w

    wh = wb.create_sheet("Human_Checklist")
    wh.append(["Symbol", "Group", "Non-Negotiable", "Item", "Done?"])
    for r in results:
        for title, non_neg, items in r["checklist"]:
            for it in items:
                wh.append([r["fin"]["sym"], title, "YES" if non_neg else "", it, ""])
                if non_neg:
                    wh.cell(wh.max_row, 3).fill = PatternFill("solid", fgColor=FILL["DANGER"])
    _style_header(wh)
    for col, w in {"A": 10, "B": 40, "C": 14, "D": 72, "E": 8}.items():
        wh.column_dimensions[col].width = w

    wc = wb.create_sheet("CrossCheck")
    wc.append(["Symbol", "Section", "Metric", "Computed", "scstrade", "Delta_%", "Note"])
    for r in results:
        for sec, metric, comp, sv, note, scale in cross_check(r["fin"]):
            delta = None
            if comp is not None and sv not in (None, 0):
                delta = round(abs(comp * scale - sv) / abs(sv) * 100, 1)
            wc.append([r["fin"]["sym"], sec, metric,
                       round(comp, 3) if isinstance(comp, float) else comp,
                       round(sv, 3) if isinstance(sv, float) else sv, delta, note])
            if delta is not None and delta > 5 and note and "scstrade uses" not in note and "ending" not in note:
                wc.cell(wc.max_row, 6).fill = PatternFill("solid", fgColor=FILL["WARN"])
    _style_header(wc)
    for col, w in {"A": 10, "B": 14, "C": 24, "D": 14, "E": 14, "F": 9, "G": 46}.items():
        wc.column_dimensions[col].width = w

    wb.save(path)
    return path


def main(argv=None):
    ap = argparse.ArgumentParser(description="PSX fundamental screener")
    ap.add_argument("symbols", nargs="+")
    ap.add_argument("--riskfree", type=float, default=DEFAULT_RISKFREE,
                    help="T-bill / policy rate %% for the earnings-yield comparison")
    ap.add_argument("--out", default="psx_report")
    args = ap.parse_args(argv)

    results = []
    for s in args.symbols:
        try:
            r = analyze(s, args.riskfree)
            results.append(r)
            print_console(r)
        except Exception as e:
            print(f"\n{'=' * 88}\n {s.upper()}  [ERROR] {e}")
    if results:
        path = write_xlsx(results, f"{args.out}.xlsx")
        import os
        print(f"\nSaved {os.path.abspath(path)}  "
              f"(sheets: Verdicts, Flags, Human_Checklist, CrossCheck)")
    print("\nScreening tool, not investment advice. No final BUY is issued until the "
          "human\ngovernance gate is confirmed.")


if __name__ == "__main__":
    main()
